import os
from openpyxl import load_workbook


# Called in process_client.py
def generate_final_report(logger):
    template_file = os.path.join("src", "template.xlsx")
    directory_path = os.path.join("output")

    # logger.info('something') prints to the console when the program runs
    # logger.debug('something') only prints to the log file

    # Create a dictionary to map the sheet names
    sheet_mapping = {
        "Unassociated Elastic IPs": "Unattached Elastic IPs",
        "Old AMIs": "EC2 Old Image",
        "Old EBS Snapshots": "EC2 Old Snapshots",
        "Unattached Volumes": "Unattached EBS Volumes",
        "Unused AMIs": "EC2 Image Not Associated",
        "Old RDS Snapshots": "RDS Old Snapshots",
        # Add more mappings as needed
    }

    # Iterate over all files in the directory
    for filename in os.listdir(directory_path):
        if "Validated" in filename and filename.endswith(".xlsx"):
            # Get the full file path
            file_path = os.path.join(directory_path, filename)

            # Load the validated Excel file
            validated_wb = load_workbook(file_path)

            # Create a new workbook based on the template
            final_wb = load_workbook(template_file)

            # Iterate over the sheets in the validated workbook
            for validated_sheet in validated_wb.sheetnames:
                if validated_sheet in sheet_mapping and sheet_mapping[validated_sheet] in final_wb.sheetnames:
                    # Get the corresponding sheet in the final workbook
                    final_sheet = final_wb[sheet_mapping[validated_sheet]]

                    # Get the corresponding sheet in the validated workbook
                    validated_sheet = validated_wb[validated_sheet]

                    # Copy the data from the validated sheet to the final sheet
                    for i, row in enumerate(validated_sheet.iter_rows(values_only=True), start=1):
                        if i > 1:  # Exclude the first row (header)
                            final_sheet.append(row)

            # Generate the new filename by replacing "Validated" with "Final"
            new_filename = filename.replace("Validated", "Final")

            # Save the final workbook with the new filename
            final_wb.save(os.path.join(directory_path, new_filename))

            # Close the workbooks
            validated_wb.close()
            final_wb.close()

    return
